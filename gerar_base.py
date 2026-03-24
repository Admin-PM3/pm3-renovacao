"""
gerar_base.py - Base de Renovacao PM3
Extrai pagamentos confirmados + certificados e gera XLSX consolidado
(uma linha por cliente). Somente leitura. Nunca executa INSERT, UPDATE ou DELETE.
"""

import os
import sys
import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# Forcar UTF-8 no stdout
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

load_dotenv()

# ─── PARAMETROS CONFIGURÁVEIS ────────────────────────────────────────────────
DATA_INICIO = "2025-01-01"   # inicio do periodo (inclusive)
DATA_FIM    = "2025-04-30"   # fim do periodo (inclusive)

# Status que representam pagamento bem-sucedido
# - succeeded: pagamento confirmado e capturado pelo gateway
# - authorized: pagamento autorizado (usado em renovacoes/assinaturas)
STATUS_SUCESSO = ("succeeded", "authorized")

# ─── Conexões ────────────────────────────────────────────────────────────────

def conectar_pagamentos():
    port = os.getenv("DB_PAYMENTS_PORT")
    if port is None:
        raise ValueError("Variável de ambiente DB_PAYMENTS_PORT não configurada.")
    return psycopg2.connect(
        host=os.getenv("DB_PAYMENTS_HOST"),
        port=int(port),
        dbname=os.getenv("DB_PAYMENTS_DB"),
        user=os.getenv("DB_PAYMENTS_USER"),
        password=os.getenv("DB_PAYMENTS_PASSWORD"),
        sslmode="require",
        connect_timeout=15,
    )

def conectar_certificados():
    port = os.getenv("DB_CERTS_PORT")
    if port is None:
        raise ValueError("Variável de ambiente DB_CERTS_PORT não configurada.")
    return psycopg2.connect(
        host=os.getenv("DB_CERTS_HOST"),
        port=int(port),
        dbname=os.getenv("DB_CERTS_DB"),
        user=os.getenv("DB_CERTS_USER"),
        password=os.getenv("DB_CERTS_PASSWORD"),
        sslmode="require",
        connect_timeout=15,
    )

# ─── ETAPA A: Extrair Pagamentos (com JOINs) ─────────────────────────────────

def extrair_pagamentos(data_inicio=None, data_fim=None):
    data_inicio = data_inicio or DATA_INICIO
    data_fim    = data_fim    or DATA_FIM
    print(f"\n[1/5] Extraindo pagamentos de {data_inicio} a {data_fim}...")
    print(f"      Status incluidos: {', '.join(STATUS_SUCESSO)}")

    conn = conectar_pagamentos()
    cur = conn.cursor()

    sql = """
        SELECT
            p.client_name,
            p.client_email,
            p.client_document,
            p.product_name,
            p.product_id,
            p.total_amount,
            p.status,
            p.created_at,
            p.subscription_id,
            p.coupon_id,
            p.coupon_percent_off,
            cp.phone_number AS telefone,
            pr.access_time
        FROM public.payments p
        LEFT JOIN public.client_payments cp ON p.client_payment_id = cp.id
        LEFT JOIN public.products pr        ON p.product_id = pr.id
        WHERE
            p.created_at >= %s::timestamp
            AND p.created_at < (%s::date + INTERVAL '1 day')::timestamp
            AND p.status IN %s
        ORDER BY p.client_email, p.created_at DESC
    """
    cur.execute(sql, (data_inicio, data_fim, STATUS_SUCESSO))

    rows = cur.fetchall()
    col_names = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    df = pd.DataFrame(rows, columns=col_names)
    print(f"      -> {len(df):,} pagamentos encontrados | {df['client_email'].nunique():,} emails unicos")
    return df

# ─── ETAPA B: Extrair Certificados ───────────────────────────────────────────

def extrair_certificados(emails_lista):
    print(f"\n[2/5] Buscando certificados para {len(emails_lista):,} emails...")

    conn = conectar_certificados()
    cur = conn.cursor()

    chunk_size = 500
    frames = []
    for i in range(0, len(emails_lista), chunk_size):
        chunk = emails_lista[i:i+chunk_size]
        placeholders = ','.join(['%s'] * len(chunk))
        sql = f"SELECT * FROM prod.cursos_certificados WHERE LOWER(TRIM(email)) IN ({placeholders})"
        cur.execute(sql, [e.lower().strip() for e in chunk])
        rows = cur.fetchall()
        if rows:
            col_names = [desc[0] for desc in cur.description]
            frames.append(pd.DataFrame(rows, columns=col_names))

    cur.close()
    conn.close()

    if frames:
        df_certs = pd.concat(frames, ignore_index=True)
    else:
        df_certs = pd.DataFrame(columns=['email','credencial_customer_id','data_de_emissao','curso','link_do_backup','nome'])

    print(f"      -> {len(df_certs):,} certificados | {df_certs['email'].nunique():,} alunos distintos")
    return df_certs


# ─── ETAPA A2: Histórico completo (sem filtro de data) ───────────────────────

def extrair_historico_completo(emails_lista):
    """Busca a compra mais recente de cada cliente SEM filtro de data.
    Retorna uma linha por cliente com ultima_compra_geral, access_time_hist e subscription_id.
    """
    print(f"\n[+] Buscando historico completo para {len(emails_lista):,} emails (sem filtro de data)...")

    conn = conectar_pagamentos()
    cur  = conn.cursor()
    frames     = []
    chunk_size = 500

    for i in range(0, len(emails_lista), chunk_size):
        chunk        = [e.lower().strip() for e in emails_lista[i:i + chunk_size]]
        placeholders = ','.join(['%s'] * len(chunk))
        sql = f"""
            SELECT DISTINCT ON (LOWER(TRIM(p.client_email)))
                LOWER(TRIM(p.client_email))  AS _email_key,
                p.created_at                 AS ultima_compra_geral,
                p.product_name               AS produto_ultima_compra,
                p.total_amount               AS valor_ultima_compra,
                p.subscription_id,
                COALESCE(pr.access_time, 0)  AS access_time_hist
            FROM public.payments p
            LEFT JOIN public.products pr ON p.product_id = pr.id
            WHERE LOWER(TRIM(p.client_email)) IN ({placeholders})
              AND p.status IN ('succeeded', 'authorized')
            ORDER BY LOWER(TRIM(p.client_email)), p.created_at DESC
        """
        cur.execute(sql, chunk)
        rows = cur.fetchall()
        if rows:
            col_names = [desc[0] for desc in cur.description]
            frames.append(pd.DataFrame(rows, columns=col_names))

    cur.close()
    conn.close()

    if frames:
        df = pd.concat(frames, ignore_index=True).drop_duplicates(subset='_email_key')
    else:
        df = pd.DataFrame(columns=[
            '_email_key', 'ultima_compra_geral', 'produto_ultima_compra',
            'valor_ultima_compra', 'subscription_id', 'access_time_hist',
        ])

    print(f"      -> {len(df):,} registros de historico encontrados")
    return df


# ─── ETAPA D: Assinaturas ────────────────────────────────────────────────────

def extrair_subscriptions(subscription_ids):
    """Busca status e próxima cobrança das assinaturas ativas."""
    ids_validos = list({
        str(s) for s in subscription_ids
        if s and str(s).strip() not in ('', 'nan', 'None', 'NaN')
    })

    if not ids_validos:
        print("\n[+] Nenhuma subscription_id encontrada. Pulando busca de assinaturas.")
        return pd.DataFrame(columns=['id', 'status', 'next_billing_at'])

    print(f"\n[+] Buscando {len(ids_validos):,} assinaturas...")

    try:
        conn = conectar_pagamentos()
        cur  = conn.cursor()
        frames     = []
        chunk_size = 500

        for i in range(0, len(ids_validos), chunk_size):
            chunk        = ids_validos[i:i + chunk_size]
            placeholders = ','.join(['%s'] * len(chunk))
            sql = f"""
                SELECT id, status, next_billing_at
                FROM public.subscriptions
                WHERE id IN ({placeholders})
            """
            cur.execute(sql, chunk)
            rows = cur.fetchall()
            if rows:
                col_names = [desc[0] for desc in cur.description]
                frames.append(pd.DataFrame(rows, columns=col_names))

        cur.close()
        conn.close()

        if frames:
            df     = pd.concat(frames, ignore_index=True).drop_duplicates(subset='id')
            ativos = (df['status'] == 'active').sum()
            print(f"      -> {len(df):,} assinaturas encontradas | {ativos:,} ativas")
            return df
        return pd.DataFrame(columns=['id', 'status', 'next_billing_at'])

    except Exception as e:
        print(f"      [AVISO] Erro ao buscar assinaturas: {e}. Continuando sem dados de assinatura.")
        return pd.DataFrame(columns=['id', 'status', 'next_billing_at'])


# ─── ETAPA C: Consolidar (uma linha por cliente) ──────────────────────────────

def consolidar(df_pag, df_certs, df_historico=None, df_subs=None, data_fim=None):
    print("\n[3/5] Consolidando dados (uma linha por cliente)...")

    hoje = datetime.today().date()

    # ── 1. Normalizar emails ───────────────────────────────────────────────────
    df_pag['_email_key']   = df_pag['client_email'].str.lower().str.strip()
    df_certs['_email_key'] = df_certs['email'].str.lower().str.strip()
    df_pag['created_at']   = pd.to_datetime(df_pag['created_at'], errors='coerce')

    # ── 2. Compra mais recente do PERÍODO (para Produto que ja comprou, etc.) ──
    df_recente = (
        df_pag
        .sort_values('created_at', ascending=False)
        .groupby('_email_key', as_index=False)
        .first()
    )

    # ── 3. Merge com histórico completo ────────────────────────────────────────
    if df_historico is not None and len(df_historico) > 0:
        df_historico['ultima_compra_geral'] = pd.to_datetime(
            df_historico['ultima_compra_geral'], errors='coerce'
        )
        hist_cols = ['_email_key', 'ultima_compra_geral', 'produto_ultima_compra',
                     'valor_ultima_compra', 'subscription_id', 'access_time_hist']
        df_recente = df_recente.merge(
            df_historico[hist_cols],
            on='_email_key', how='left', suffixes=('', '_hist')
        )
        # Preferir subscription_id do historico (compra mais recente geral)
        if 'subscription_id_hist' in df_recente.columns:
            df_recente['subscription_id'] = (
                df_recente['subscription_id_hist']
                .fillna(df_recente['subscription_id'])
            )
            df_recente.drop(columns=['subscription_id_hist'], inplace=True)
    else:
        # Fallback: usar dados do período
        df_recente['ultima_compra_geral']   = df_recente['created_at']
        df_recente['produto_ultima_compra'] = df_recente['product_name']
        df_recente['valor_ultima_compra']   = df_recente['total_amount']
        df_recente['access_time_hist']      = df_recente.get('access_time', pd.Series(dtype=object))

    # ── 4. Flag: comprou_apos_periodo ──────────────────────────────────────────
    if data_fim:
        try:
            data_fim_date = datetime.strptime(data_fim, '%Y-%m-%d').date()
            df_recente['comprou_apos_periodo'] = df_recente['ultima_compra_geral'].apply(
                lambda v: v.date() > data_fim_date if pd.notna(v) else False
            )
        except Exception:
            df_recente['comprou_apos_periodo'] = False
    else:
        df_recente['comprou_apos_periodo'] = False

    # ── 5. Data de renovação baseada em ultima_compra_geral ────────────────────
    def calc_renovacao(row):
        try:
            meses = int(row['access_time_hist'])
            if meses > 0 and pd.notna(row['ultima_compra_geral']):
                base = row['ultima_compra_geral']
                base = base.date() if hasattr(base, 'date') else base
                return (base + relativedelta(months=meses)).strftime('%d/%m/%Y')
        except Exception:
            pass
        return ""

    df_recente['Data da possivel renovacao'] = df_recente.apply(calc_renovacao, axis=1)

    # ── 6. Formatar total_amount (centavos → R$) ───────────────────────────────
    def fmt_valor(v):
        try:
            return f"R$ {float(v)/100:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except Exception:
            return str(v) if v is not None else ""

    df_recente['total_amount_fmt'] = df_recente['total_amount'].apply(fmt_valor)

    # ── 7. Dias desde última compra ────────────────────────────────────────────
    def dias_desde_compra(v):
        try:
            d = pd.to_datetime(v)
            if pd.notna(d):
                return (hoje - d.date()).days
        except Exception:
            pass
        return None

    df_recente['dias_desde_ultima_compra'] = df_recente['ultima_compra_geral'].apply(dias_desde_compra)

    # ── 8. Agregar certificados por email ──────────────────────────────────────
    certs_por_email = (
        df_certs
        .sort_values('data_de_emissao', ascending=False)
        .groupby('_email_key', as_index=False)
        .agg(
            quant_certs=('curso', 'count'),
            quais_certs=('curso', lambda x: ' | '.join(x.dropna().unique())),
            data_cert_recente=('data_de_emissao', 'first'),
        )
    )

    df_merge = df_recente.merge(certs_por_email, on='_email_key', how='left')

    # Formatar data do certificado
    def fmt_data(v):
        if pd.isna(v) or v is None or v == "":
            return ""
        try:
            return pd.to_datetime(v, dayfirst=True).strftime('%d/%m/%Y')
        except Exception:
            return str(v)

    df_merge['data_cert_recente'] = df_merge['data_cert_recente'].apply(fmt_data)

    # ── 9. Dias desde último certificado ──────────────────────────────────────
    def dias_cert(v):
        if not v or v == "":
            return None
        try:
            return (hoje - datetime.strptime(v, '%d/%m/%Y').date()).days
        except Exception:
            return None

    df_merge['dias_desde_ultimo_certificado'] = df_merge['data_cert_recente'].apply(dias_cert)

    # ── 10. Merge com assinaturas ──────────────────────────────────────────────
    if df_subs is not None and len(df_subs) > 0:
        df_subs_m                   = df_subs.copy()
        df_subs_m['_sub_id']        = df_subs_m['id'].astype(str)
        df_subs_m['status_assinatura'] = df_subs_m['status'].fillna('')
        df_subs_m['proxima_cobranca']  = df_subs_m['next_billing_at'].apply(
            lambda v: pd.to_datetime(v).strftime('%d/%m/%Y')
            if pd.notna(v) and str(v).strip() not in ('', 'None', 'nan') else ''
        )
        df_subs_m = df_subs_m[['_sub_id', 'status_assinatura', 'proxima_cobranca']]

        df_merge['_sub_id'] = df_merge['subscription_id'].fillna('').astype(str)
        df_merge = df_merge.merge(df_subs_m, on='_sub_id', how='left')
        df_merge.drop(columns=['_sub_id'], inplace=True)
        df_merge['possui_assinatura_ativa'] = df_merge['status_assinatura'] == 'active'
        df_merge['status_assinatura']  = df_merge['status_assinatura'].fillna('')
        df_merge['proxima_cobranca']   = df_merge['proxima_cobranca'].fillna('')
    else:
        df_merge['possui_assinatura_ativa'] = False
        df_merge['status_assinatura']       = ''
        df_merge['proxima_cobranca']        = ''

    # ── 11. Tipo de cliente ────────────────────────────────────────────────────
    def classificar_tipo(row):
        if row.get('possui_assinatura_ativa') is True:
            return 'Assinatura ativa'
        if row.get('comprou_apos_periodo') is True:
            return 'Ja renovou'
        try:
            if pd.notna(row.get('quant_certs')) and int(row.get('quant_certs', 0)) > 0:
                return 'Elegivel para renovacao'
        except Exception:
            pass
        return 'Sem engajamento'

    df_merge['tipo_cliente'] = df_merge.apply(classificar_tipo, axis=1)

    # ── 12. Formatar ultima_compra_geral ───────────────────────────────────────
    df_merge['ultima_compra_fmt'] = df_merge['ultima_compra_geral'].apply(
        lambda v: v.strftime('%d/%m/%Y') if pd.notna(v) else ""
    )

    # ── 13. Separar qualificados e sem certificado ─────────────────────────────
    mask_com_cert = df_merge['quant_certs'].notna() & (df_merge['quant_certs'] > 0)
    df_com = df_merge[mask_com_cert].copy()
    df_sem = df_merge[~mask_com_cert].copy()

    # ── 14. Construir colunas finais ───────────────────────────────────────────
    def montar_df_final(df, tem_cert=True):
        out = pd.DataFrame()
        # ── Colunas originais (9) ──
        out['Nome']                          = df['client_name']
        out['Email']                         = df['_email_key']
        out['Telefone']                      = df['telefone'].fillna('')
        out['Produto que ja comprou']        = df['product_name'].fillna('')
        out['Valor que pagou']               = df['total_amount_fmt']
        out['Data da possivel renovacao']    = df['Data da possivel renovacao']
        if tem_cert:
            out['Quantos certificados emitidos'] = df['quant_certs'].fillna(0).astype(int)
            out['Quais certificados emitidos']   = df['quais_certs'].fillna('')
            out['Data de Emissao de certificado']= df['data_cert_recente'].fillna('')
        else:
            out['Quantos certificados emitidos'] = ''
            out['Quais certificados emitidos']   = ''
            out['Data de Emissao de certificado']= ''
        # ── Novas colunas ──
        out['Ultima compra (geral)']         = df['ultima_compra_fmt']
        out['Comprou apos periodo']          = df['comprou_apos_periodo'].map(
            {True: 'Sim', False: 'Nao'}
        ).fillna('Nao')
        out['Possui assinatura ativa']       = df['possui_assinatura_ativa'].map(
            {True: 'Sim', False: 'Nao'}
        ).fillna('Nao')
        out['Status assinatura']             = df['status_assinatura'].fillna('')
        out['Proxima cobranca']              = df['proxima_cobranca'].fillna('')
        out['Dias desde ultima compra']      = df['dias_desde_ultima_compra'].apply(
            lambda v: int(v) if pd.notna(v) else ''
        )
        out['Dias desde ultimo certificado'] = df['dias_desde_ultimo_certificado'].apply(
            lambda v: int(v) if pd.notna(v) else ''
        )
        out['Tipo de cliente']               = df['tipo_cliente'].fillna('Sem engajamento')
        return out.reset_index(drop=True)

    df_qual_final = montar_df_final(df_com, tem_cert=True)
    df_sem_final  = montar_df_final(df_sem, tem_cert=False)

    # ── 15. Validações finais ──────────────────────────────────────────────────
    dup_qual = df_qual_final['Email'].duplicated().sum()
    dup_sem  = df_sem_final['Email'].duplicated().sum()
    assert dup_qual == 0, f"ERRO: {dup_qual} emails duplicados em Base Qualificada!"
    assert dup_sem  == 0, f"ERRO: {dup_sem} emails duplicados em Sem Certificado!"

    novas_cols = [
        'Ultima compra (geral)', 'Comprou apos periodo', 'Possui assinatura ativa',
        'Status assinatura', 'Proxima cobranca', 'Dias desde ultima compra',
        'Dias desde ultimo certificado', 'Tipo de cliente',
    ]
    for col in novas_cols:
        assert col in df_qual_final.columns, f"ERRO: coluna '{col}' ausente no output!"

    print(f"      -> Qualificados: {len(df_qual_final):,} clientes (com certificado)")
    print(f"      -> Sem certificado: {len(df_sem_final):,} clientes")
    print(f"      -> Total colunas no output: {len(df_qual_final.columns)}")

    return df_qual_final, df_sem_final, df_recente

# ─── ETAPA D: Calcular métricas de renovacao ─────────────────────────────────

def calcular_renovacoes(df_qual):
    """Conta renovacoes nos proximos 30 e 90 dias a partir de hoje."""
    hoje = date.today()
    def parse_dt(v):
        try:
            return datetime.strptime(v, '%d/%m/%Y').date()
        except Exception:
            return None

    datas = df_qual['Data da possivel renovacao'].apply(parse_dt)
    prox30 = sum(1 for d in datas if d and hoje <= d <= hoje + relativedelta(days=30))
    prox90 = sum(1 for d in datas if d and hoje <= d <= hoje + relativedelta(days=90))
    return prox30, prox90

# ─── ETAPA E: Gerar XLSX ─────────────────────────────────────────────────────

def formatar_aba(ws, df, zebra=True, destacar_renovacao=False):
    """Aplica formatacao: cabecalho azul, zebra, freeze, autowidth, filtro."""
    AZUL_ESCURO = "1a3a5c"
    CINZA_CLARO = "f5f5f5"
    AMARELO     = "fff3cd"
    BRANCO      = "FFFFFF"

    header_fill  = PatternFill("solid", fgColor=AZUL_ESCURO)
    header_font  = Font(color=BRANCO, bold=True)
    gray_fill    = PatternFill("solid", fgColor=CINZA_CLARO)
    yellow_fill  = PatternFill("solid", fgColor=AMARELO)

    hoje = date.today()
    limite90 = hoje + relativedelta(days=90)

    # Indice da coluna de renovacao (0-based)
    col_renov_idx = None
    if destacar_renovacao and 'Data da possivel renovacao' in df.columns:
        col_renov_idx = list(df.columns).index('Data da possivel renovacao')

    # Cabecalho
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            if not isinstance(val, (list, dict)):
                try:
                    if pd.isna(val):
                        val = ""
                except Exception:
                    pass
            if hasattr(val, 'item'):
                val = val.item()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            # Zebra padrao
            if zebra and row_idx % 2 == 0:
                cell.fill = gray_fill

            # Destacar renovacao em amarelo se dentro de 90 dias
            if destacar_renovacao and col_renov_idx is not None and (col_idx - 1) == col_renov_idx:
                try:
                    d = datetime.strptime(str(val), '%d/%m/%Y').date()
                    if hoje <= d <= limite90:
                        cell.fill = yellow_fill
                except Exception:
                    pass

    # Largura automatica
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx-1].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 70)

    ws.freeze_panes = ws['A2']
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


def gerar_resumo(ws, df_recente, df_qual, df_sem, prox30, prox90, status_incluidos, data_ini, data_fim):
    AZUL_ESCURO = "1a3a5c"
    CINZA_MED   = "dce6f1"
    BRANCO      = "FFFFFF"

    header_fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    header_font = Font(color=BRANCO, bold=True)
    alt_fill    = PatternFill("solid", fgColor=CINZA_MED)

    total_clientes   = len(df_recente)
    alunos_com_cert  = len(df_qual)
    alunos_sem_cert  = len(df_sem)
    pct              = (alunos_com_cert / total_clientes * 100) if total_clientes > 0 else 0

    # Top cursos
    if 'Quais certificados emitidos' in df_qual.columns and len(df_qual) > 0:
        from collections import Counter
        todos_cursos = []
        for val in df_qual['Quais certificados emitidos']:
            if val:
                todos_cursos.extend([c.strip() for c in str(val).split('|') if c.strip()])
        top_cursos_raw = Counter(todos_cursos).most_common(10)
        top_cursos = pd.DataFrame(top_cursos_raw, columns=['curso', 'count'])
    else:
        top_cursos = pd.DataFrame(columns=['curso', 'count'])

    # Top produtos (usando df_recente que tem product_name)
    top_produtos = df_recente['product_name'].value_counts().head(10).reset_index()
    top_produtos.columns = ['produto', 'count']

    curso_top   = top_cursos.iloc[0]['curso'] if len(top_cursos) > 0 else '-'
    curso_top_n = int(top_cursos.iloc[0]['count']) if len(top_cursos) > 0 else 0
    prod_top    = top_produtos.iloc[0]['produto'] if len(top_produtos) > 0 else '-'
    prod_top_n  = int(top_produtos.iloc[0]['count']) if len(top_produtos) > 0 else 0

    d_ini_fmt = datetime.strptime(data_ini, '%Y-%m-%d').strftime('%d/%m/%Y')
    d_fim_fmt = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')

    metricas = [
        ("Metrica", "Valor"),
        ("Periodo analisado", f"{d_ini_fmt} a {d_fim_fmt}"),
        ("Status incluidos na analise", ', '.join(status_incluidos)),
        ("Total de clientes unicos no periodo", total_clientes),
        ("Alunos COM certificado (qualificados)", alunos_com_cert),
        ("Alunos SEM certificado", alunos_sem_cert),
        ("% qualificados sobre total", f"{pct:.1f}%"),
        ("Renovacoes previstas nos proximos 30 dias", prox30),
        ("Renovacoes previstas nos proximos 90 dias", prox90),
        ("Curso mais concluido", f"{curso_top} ({curso_top_n} vezes)"),
        ("Produto mais comprado", f"{prod_top} ({prod_top_n} vezes)"),
    ]

    for row_idx, (met, val) in enumerate(metricas, 1):
        c1 = ws.cell(row=row_idx, column=1, value=met)
        c2 = ws.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            c1.fill = header_fill; c1.font = header_font
            c2.fill = header_fill; c2.font = header_font
        elif row_idx % 2 == 0:
            c1.fill = alt_fill; c2.fill = alt_fill

    # Top 10 cursos
    row_start = len(metricas) + 2
    ws.cell(row=row_start, column=1, value="Top 10 Cursos mais Concluidos").fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.cell(row=row_start, column=1).font = header_font
    ws.cell(row=row_start, column=2, value="Certificados").fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.cell(row=row_start, column=2).font = header_font
    for i, row in enumerate(top_cursos.itertuples(index=False), row_start + 1):
        ws.cell(row=i, column=1, value=row.curso)
        ws.cell(row=i, column=2, value=int(row.count))
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = alt_fill
            ws.cell(row=i, column=2).fill = alt_fill

    # Top 10 produtos
    row_start2 = row_start + len(top_cursos) + 2
    ws.cell(row=row_start2, column=1, value="Top 10 Produtos mais Comprados").fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.cell(row=row_start2, column=1).font = header_font
    ws.cell(row=row_start2, column=2, value="Compras").fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.cell(row=row_start2, column=2).font = header_font
    for i, row in enumerate(top_produtos.itertuples(index=False), row_start2 + 1):
        ws.cell(row=i, column=1, value=row.produto)
        ws.cell(row=i, column=2, value=int(row.count))
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = alt_fill
            ws.cell(row=i, column=2).fill = alt_fill

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 55
    ws.freeze_panes = ws['A2']

    return {
        'total_clientes': total_clientes,
        'alunos_com_cert': alunos_com_cert,
        'alunos_sem_cert': alunos_sem_cert,
        'pct': pct,
        'prox30': prox30,
        'prox90': prox90,
        'top_cursos': top_cursos,
        'top_produtos': top_produtos,
    }


def salvar_xlsx(df_qual, df_sem, df_recente, prox30, prox90):
    print("\n[4/5] Gerando arquivo XLSX...")

    d_ini = DATA_INICIO.replace('-', '')
    d_fim = DATA_FIM.replace('-', '')
    nome_arquivo = f"base_renovacao_{d_ini}_{d_fim}.xlsx"

    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        df_qual.to_excel(writer, sheet_name='Base Qualificada', index=False)
        df_sem.to_excel(writer,  sheet_name='Sem Certificado',  index=False)
        pd.DataFrame().to_excel(writer, sheet_name='Resumo', index=False)

    wb = load_workbook(nome_arquivo)

    formatar_aba(wb['Base Qualificada'], df_qual, zebra=True, destacar_renovacao=True)
    formatar_aba(wb['Sem Certificado'],  df_sem,  zebra=True, destacar_renovacao=False)

    metricas = gerar_resumo(
        wb['Resumo'], df_recente, df_qual, df_sem,
        prox30, prox90,
        list(STATUS_SUCESSO), DATA_INICIO, DATA_FIM
    )

    wb.save(nome_arquivo)
    print(f"      -> Arquivo salvo: {nome_arquivo}")
    return nome_arquivo, metricas

# ─── IMPRESSÃO FINAL ──────────────────────────────────────────────────────────

def imprimir_resultado(nome_arquivo, metricas):
    top3_cursos = metricas['top_cursos'].head(3)
    L = chr(9553)  # borda lateral ║

    TL = chr(9556)   # ╔
    TR = chr(9559)   # ╗
    BL = chr(9562)   # ╚
    BR = chr(9565)   # ╝
    ML = chr(9568)   # ╠
    MR = chr(9571)   # ╣
    H  = chr(9552)   # ═

    print("\n" + TL + H*58 + TR)
    print(L + "       BASE DE RENOVACAO PM3 - GERADA COM SUCESSO       " + L)
    print(ML + H*58 + MR)
    print(L + "  CLIENTES" + " "*49 + L)
    print(L + f"    Total unico no periodo:    {metricas['total_clientes']:>6,}" + " "*22 + L)
    print(L + f"    Qualificados (com cert):   {metricas['alunos_com_cert']:>6,}  ({metricas['pct']:.1f}%)" + " "*13 + L)
    print(L + f"    Sem certificado:           {metricas['alunos_sem_cert']:>6,}" + " "*22 + L)
    print(ML + H*58 + MR)
    print(L + "  RENOVACOES PREVISTAS" + " "*37 + L)
    print(L + f"    Proximos 30 dias:  {metricas['prox30']:>5,} clientes" + " "*26 + L)
    print(L + f"    Proximos 90 dias:  {metricas['prox90']:>5,} clientes" + " "*26 + L)
    print(ML + H*58 + MR)
    print(L + "  TOP 3 CURSOS MAIS CONCLUIDOS" + " "*29 + L)
    for i, row in enumerate(top3_cursos.itertuples(index=False), 1):
        curso_trunc = str(row.curso)[:42]
        print(L + f"    {i}. {curso_trunc:<42} {int(row.count):>4}  " + L)
    print(BL + H*58 + BR)
    print()

# ─── API PARA USO WEB ─────────────────────────────────────────────────────────

def run_pipeline(data_inicio, data_fim):
    """Executa o pipeline completo e retorna os DataFrames consolidados.
    Usado pela interface web (app.py).
    Retorna: (df_qual, df_sem, df_recente, prox30, prox90)
    """
    # [1] Pagamentos do período
    df_pag = extrair_pagamentos(data_inicio, data_fim)
    if len(df_pag) == 0:
        empty_cols = [
            'Nome', 'Email', 'Telefone', 'Produto que ja comprou',
            'Valor que pagou', 'Data da possivel renovacao',
            'Quantos certificados emitidos', 'Quais certificados emitidos',
            'Data de Emissao de certificado',
            'Ultima compra (geral)', 'Comprou apos periodo',
            'Possui assinatura ativa', 'Status assinatura', 'Proxima cobranca',
            'Dias desde ultima compra', 'Dias desde ultimo certificado',
            'Tipo de cliente',
        ]
        empty = pd.DataFrame(columns=empty_cols)
        return empty, empty.copy(), pd.DataFrame(), 0, 0

    emails_lista = (
        df_pag['client_email'].dropna()
        .str.lower().str.strip()
        .unique().tolist()
    )

    # [2] Histórico completo (sem filtro de data)
    df_historico = extrair_historico_completo(emails_lista)

    # [3] Certificados
    df_certs = extrair_certificados(emails_lista)

    # [4] Assinaturas — prioriza subscription_id do histórico (mais recente)
    sub_ids = []
    if len(df_historico) > 0:
        sub_ids = df_historico['subscription_id'].dropna().unique().tolist()
    if not sub_ids:
        sub_ids = df_pag['subscription_id'].dropna().unique().tolist()
    df_subs = extrair_subscriptions(sub_ids)

    # [5] Consolidar com todos os novos dados
    df_qual, df_sem, df_recente = consolidar(
        df_pag, df_certs, df_historico, df_subs, data_fim
    )
    prox30, prox90 = calcular_renovacoes(df_qual)
    return df_qual, df_sem, df_recente, prox30, prox90


def build_xlsx_bytes(df_qual, df_sem, df_recente, prox30, prox90, data_inicio, data_fim):
    """Gera o XLSX em memória e retorna um BytesIO pronto para download."""
    import io as _io

    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_qual.to_excel(writer, sheet_name='Base Qualificada', index=False)
        df_sem.to_excel(writer,  sheet_name='Sem Certificado',  index=False)
        pd.DataFrame().to_excel(writer, sheet_name='Resumo', index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    formatar_aba(wb['Base Qualificada'], df_qual, zebra=True, destacar_renovacao=True)
    formatar_aba(wb['Sem Certificado'],  df_sem,  zebra=True, destacar_renovacao=False)
    gerar_resumo(
        wb['Resumo'], df_recente, df_qual, df_sem,
        prox30, prox90, list(STATUS_SUCESSO), data_inicio, data_fim
    )

    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "="*60)
    print("  BASE DE RENOVACAO PM3")
    print(f"  Periodo: {DATA_INICIO} a {DATA_FIM}")
    print(f"  Status sucesso: {', '.join(STATUS_SUCESSO)}")
    print("  Justificativa: 'succeeded'=pagamento capturado pelo gateway;")
    print("  'authorized'=pagamento autorizado (renovacoes/assinaturas).")
    print("  Excluidos: error, pending, failed, expired e similares.")
    print("="*60)

    try:
        # Etapa A
        df_pag = extrair_pagamentos()

        if len(df_pag) == 0:
            print("\n[AVISO] Nenhum pagamento encontrado no periodo informado.")
            print("  Verifique DATA_INICIO e DATA_FIM no topo do arquivo.")
            sys.exit(0)

        # Etapa B — historico completo
        emails_lista = (
            df_pag['client_email'].dropna()
            .str.lower().str.strip().unique().tolist()
        )
        df_historico = extrair_historico_completo(emails_lista)

        # Etapa C — certificados
        df_certs = extrair_certificados(emails_lista)

        # Etapa D — assinaturas
        sub_ids = df_historico['subscription_id'].dropna().unique().tolist() if len(df_historico) > 0 \
                  else df_pag['subscription_id'].dropna().unique().tolist()
        df_subs = extrair_subscriptions(sub_ids)

        # Etapa E — consolidar
        df_qual, df_sem, df_recente = consolidar(df_pag, df_certs, df_historico, df_subs, DATA_FIM)

        # Metricas de renovacao
        prox30, prox90 = calcular_renovacoes(df_qual)
        print(f"\n      Renovacoes proximos 30 dias: {prox30} | 90 dias: {prox90}")

        # Etapa D+E
        nome_arquivo, metricas = salvar_xlsx(df_qual, df_sem, df_recente, prox30, prox90)

        # Validacao rapida
        assert list(df_qual.columns) == [
            'Nome', 'Email', 'Telefone', 'Produto que ja comprou',
            'Valor que pagou', 'Data da possivel renovacao',
            'Quantos certificados emitidos', 'Quais certificados emitidos',
            'Data de Emissao de certificado'
        ], "ERRO: colunas de Base Qualificada incorretas!"

        print("\n[5/5] Concluido!")
        imprimir_resultado(nome_arquivo, metricas)

    except Exception as e:
        print(f"\n[ERRO] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
